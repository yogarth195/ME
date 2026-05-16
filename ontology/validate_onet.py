"""
Validate ontology coverage against O*NET Technology Skills data.

O*NET publishes a Technology Skills dataset (Excel) at:
  https://www.onetcenter.org/dl_files/database/db_29_0_excel/Technology%20Skills.xlsx

This script:
1. Downloads the Excel file (or uses a cached copy)
2. Extracts the "Example" column (technology tool names)
3. Cross-references against our 266 skill nodes
4. Reports: coverage %, matched nodes, and unmatched O*NET tools

Usage:
    python ontology/validate_onet.py
    python ontology/validate_onet.py --cached   # skip download, use cached xlsx
"""

import argparse
import pickle
import re
import urllib.request
from pathlib import Path

ONTOLOGY_DIR = Path(__file__).parent
CACHE_PATH = ONTOLOGY_DIR / "_onet_technology_skills.xlsx"
ONET_URL = (
    "https://www.onetcenter.org/dl_files/database/db_29_0_excel/"
    "Technology%20Skills.xlsx"
)


def _load_skill_nodes() -> set:
    graph_path = ONTOLOGY_DIR / "skill_ontology.gpickle"
    with open(graph_path, "rb") as f:
        G = pickle.load(f)
    return {n for n in G.nodes if not n.startswith("_alias_")}


def _download_onet(path: Path) -> None:
    print(f"Downloading O*NET Technology Skills -> {path.name} ...")
    headers = {"User-Agent": "ExplainHire-OntologyValidator/1.0"}
    req = urllib.request.Request(ONET_URL, headers=headers)
    with urllib.request.urlopen(req, timeout=30) as resp, open(path, "wb") as f:
        f.write(resp.read())
    print(f"Downloaded ({path.stat().st_size // 1024} KB)")


def _parse_onet_tools(path: Path) -> set:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "openpyxl required to parse the xlsx file.\n"
            "Install it:  pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    example_col = next((i for i, h in enumerate(headers) if "example" in h), None)
    if example_col is None:
        raise ValueError(f"Could not find 'Example' column. Headers found: {headers}")

    tools = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[example_col]
        if val:
            # Normalise: lowercase, replace spaces/hyphens with underscore
            norm = re.sub(r"[\s\-/]+", "_", str(val).strip().lower())
            norm = re.sub(r"[^a-z0-9_]", "", norm)
            tools.add(norm)

    return tools


def _fuzzy_match(onet_tool: str, skill_nodes: set) -> str | None:
    """Check if an O*NET tool name is contained in or contains a skill node."""
    if onet_tool in skill_nodes:
        return onet_tool
    for node in skill_nodes:
        if node in onet_tool or onet_tool in node:
            return node
    return None


def validate(use_cache: bool = False) -> dict:
    skill_nodes = _load_skill_nodes()
    print(f"Ontology skill nodes: {len(skill_nodes)}")

    if not use_cache or not CACHE_PATH.exists():
        _download_onet(CACHE_PATH)
    else:
        print(f"Using cached file: {CACHE_PATH.name}")

    onet_tools = _parse_onet_tools(CACHE_PATH)
    print(f"O*NET technology tools (normalised): {len(onet_tools)}")

    exact_matches = skill_nodes & onet_tools
    fuzzy_matches = {}
    for tool in onet_tools - exact_matches:
        matched_node = _fuzzy_match(tool, skill_nodes)
        if matched_node:
            fuzzy_matches[tool] = matched_node

    total_matched = len(exact_matches) + len(set(fuzzy_matches.values()) - exact_matches)
    coverage = round(total_matched / len(skill_nodes) * 100, 1)

    # O*NET tools not covered at all
    all_matched_onet = exact_matches | set(fuzzy_matches.keys())
    unmatched_onet = sorted(onet_tools - all_matched_onet)[:50]  # top 50 gaps

    print(f"\n=== O*NET Coverage Report ===")
    print(f"Exact matches   : {len(exact_matches)} / {len(skill_nodes)} nodes ({len(exact_matches)/len(skill_nodes)*100:.1f}%)")
    print(f"Fuzzy matches   : {len(fuzzy_matches)} additional nodes")
    print(f"Total coverage  : {total_matched} / {len(skill_nodes)} = {coverage}%")
    print(f"\nTop O*NET tools NOT in our ontology (sample):")
    for tool in unmatched_onet[:20]:
        print(f"  - {tool}")

    return {
        "skill_nodes":     len(skill_nodes),
        "onet_tools":      len(onet_tools),
        "exact_matches":   sorted(exact_matches),
        "fuzzy_matches":   fuzzy_matches,
        "coverage_pct":    coverage,
        "unmatched_onet":  unmatched_onet,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--cached", action="store_true", help="Use cached xlsx if available")
    args = parser.parse_args()
    validate(use_cache=args.cached)
