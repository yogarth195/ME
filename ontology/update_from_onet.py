"""
Update skill ontology with real tech skills from O*NET Technology Skills data.

Source: data/raw/onet_tech_skills.xlsx
  Columns: O*NET-SOC Code, Title, Example, Commodity Code, Commodity Title, Hot Technology, In Demand

Filters to tech occupations only, extracts tool/skill names, maps to ontology
parent categories, and adds missing nodes + edges to the graph.

Run:
    python ontology/update_from_onet.py

Then rebuild ontology:
    python ontology/build_ontology.py
"""

import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

XLSX_PATH = Path(__file__).parent.parent / "data" / "raw" / "onet_tech_skills.xlsx"
ONTOLOGY_PATH = Path(__file__).parent / "build_ontology.py"

# ── Tech occupation keywords (filter non-tech roles out) ─────────────────────
TECH_TITLE_KEYWORDS = [
    "software", "developer", "engineer", "data", "analyst", "scientist",
    "programmer", "architect", "devops", "database", "network", "security",
    "cloud", "machine learning", "artificial intelligence", "web", "mobile",
    "backend", "frontend", "full stack", "systems", "it ", "information technology",
    "computer", "technology", "platform", "infrastructure", "cyber",
]

# ── Map O*NET Commodity Title → our ontology parent node ─────────────────────
COMMODITY_TO_PARENT = {
    "development environment software":         "programming",
    "object or component oriented development": "programming",
    "program testing software":                 "testing",
    "web platform development software":        "web_development",
    "application server software":              "backend",
    "database user interface and query software": "database",
    "data base management system software":     "database",
    "data base reporting software":             "data_analysis",
    "operating system software":                "linux",
    "network monitoring software":              "networking",
    "network security or virtual private network vpn management software": "security",
    "firewall software":                        "security",
    "cloud-based data access and sharing software": "cloud_computing",
    "cloud-based management software":          "cloud_computing",
    "cloud-based protection or security software": "security",
    "configuration management software":        "devops",
    "platform interconnectivity software":      "devops",
    "content workflow software":                "project_management",
    "project management software":              "project_management",
    "customer relationship management crm software": "project_management",
    "enterprise resource planning erp software": "project_management",
    "data mining software":                     "machine_learning",
    "analytical or scientific software":        "data_analysis",
    "business intelligence and data analysis software": "data_analysis",
    "information retrieval or search software": "information_retrieval",
    "transaction server software":              "backend",
    "enterprise application integration software": "backend",
    "file versioning software":                 "git",
    "revision control software":                "git",
    "desktop communications software":          "communication",
    "video conferencing software":              "communication",
    "container software":                       "docker",
    "virtualization software":                  "virtualization",
}

# ── Skills we explicitly want to add with their parents ──────────────────────
EXPLICIT_ADDITIONS = [
    # Project management
    ("jira",            "project_management"),
    ("confluence",      "project_management"),
    ("trello",          "project_management"),
    ("asana",           "project_management"),
    ("agile",           "project_management"),
    ("scrum",           "project_management"),
    ("kanban",          "project_management"),
    ("jira_software",   "jira"),

    # Version control
    ("github",          "git"),
    ("gitlab",          "git"),
    ("bitbucket",       "git"),

    # CI/CD
    ("jenkins",         "ci_cd"),
    ("github_actions",  "ci_cd"),
    ("gitlab_ci",       "ci_cd"),
    ("circleci",        "ci_cd"),
    ("travis_ci",       "ci_cd"),
    ("ci_cd",           "devops"),

    # Containers & orchestration
    ("kubernetes",      "docker"),
    ("helm",            "kubernetes"),
    ("docker_compose",  "docker"),

    # Cloud
    ("aws",             "cloud_computing"),
    ("gcp",             "cloud_computing"),
    ("azure",           "cloud_computing"),
    ("lambda",          "aws"),
    ("s3",              "aws"),
    ("ec2",             "aws"),
    ("bigquery",        "gcp"),

    # Databases
    ("redis",           "database"),
    ("mongodb",         "nosql"),
    ("cassandra",       "nosql"),
    ("elasticsearch",   "nosql"),
    ("nosql",           "database"),
    ("postgresql",      "sql"),
    ("mysql",           "sql"),
    ("sqlite",          "sql"),
    ("sql",             "database"),

    # Messaging
    ("kafka",           "distributed_systems"),
    ("rabbitmq",        "distributed_systems"),
    ("celery",          "distributed_systems"),
    ("distributed_systems", "backend"),

    # ML/AI tools
    ("huggingface",     "deep_learning"),
    ("transformers",    "nlp"),
    ("langchain",       "nlp"),
    ("openai_api",      "nlp"),
    ("xgboost",         "machine_learning"),
    ("lightgbm",        "machine_learning"),
    ("mlflow",          "machine_learning"),
    ("airflow",         "data_engineering"),
    ("dbt",             "data_engineering"),
    ("spark",           "data_engineering"),
    ("hadoop",          "data_engineering"),
    ("data_engineering","data_science"),

    # IaC
    ("terraform",       "devops"),
    ("ansible",         "devops"),
    ("pulumi",          "devops"),

    # Monitoring
    ("grafana",         "monitoring"),
    ("prometheus",      "monitoring"),
    ("datadog",         "monitoring"),
    ("monitoring",      "devops"),

    # Frontend
    ("nextjs",          "react"),
    ("vuejs",           "javascript"),
    ("angular",         "javascript"),
    ("webpack",         "javascript"),
    ("tailwindcss",     "css"),

    # Mobile
    ("flutter",         "mobile_development"),
    ("react_native",    "mobile_development"),
    ("swift",           "mobile_development"),
    ("kotlin",          "mobile_development"),

    # Architecture / design
    ("system_design",   "software_engineering"),
    ("oop",             "software_engineering"),
    ("design_patterns", "software_engineering"),
    ("microservices",   "software_architecture"),
    ("rest_api",        "web_development"),
    ("graphql",         "web_development"),
    ("grpc",            "backend"),
    ("software_architecture", "software_engineering"),
    ("software_engineering",  "programming"),

    # Testing
    ("jest",            "testing"),
    ("pytest",          "testing"),
    ("selenium",        "testing"),
    ("cypress",         "testing"),
    ("unit_testing",    "testing"),
    ("integration_testing", "testing"),
]


def _normalise(name: str) -> str:
    name = name.lower().strip()
    name = re.sub(r"[^a-z0-9]+", "_", name)
    name = name.strip("_")
    return name


def _is_tech_occupation(title: str) -> bool:
    t = title.lower()
    return any(kw in t for kw in TECH_TITLE_KEYWORDS)


def extract_onet_skills(xlsx_path: Path) -> list[tuple[str, str]]:
    """Returns list of (skill_node, parent_node) from O*NET xlsx."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    pairs: list[tuple[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        title         = str(row[col["Title"]] or "")
        example       = str(row[col["Example"]] or "").strip()
        commodity     = str(row[col["Commodity Title"]] or "").lower().strip()
        hot           = str(row[col.get("Hot Technology", -1)] or "")
        in_demand     = str(row[col.get("In Demand", -1)] or "")

        if not _is_tech_occupation(title):
            continue
        if not example or example.lower() in ("none", "n/a", ""):
            continue

        # Only take hot or in-demand technologies to keep ontology focused
        if hot != "Y" and in_demand != "Y":
            continue

        skill_node = _normalise(example)
        if len(skill_node) < 2 or skill_node.isdigit():
            continue

        parent = None
        for key, p in COMMODITY_TO_PARENT.items():
            if key in commodity:
                parent = p
                break

        if parent:
            pairs.append((skill_node, parent))

    return pairs


def generate_additions(pairs: list[tuple[str, str]]) -> list[tuple[str, str]]:
    """Merge O*NET pairs + explicit additions, deduplicate."""
    seen: set[str] = set()
    result: list[tuple[str, str]] = []

    all_pairs = EXPLICIT_ADDITIONS + pairs

    for skill, parent in all_pairs:
        skill  = _normalise(skill)
        parent = _normalise(parent)
        if skill not in seen and skill != parent:
            seen.add(skill)
            result.append((skill, parent))

    return result


def print_report(additions: list[tuple[str, str]]):
    by_parent: dict[str, list[str]] = {}
    for skill, parent in additions:
        by_parent.setdefault(parent, []).append(skill)

    print(f"\nTotal new skill nodes to add: {len(additions)}")
    print(f"Parent categories used: {len(by_parent)}\n")
    for parent, skills in sorted(by_parent.items()):
        print(f"  {parent} ({len(skills)})")
        for s in skills[:5]:
            print(f"    - {s}")
        if len(skills) > 5:
            print(f"    ... and {len(skills)-5} more")


def write_additions_file(additions: list[tuple[str, str]]):
    """Write additions to ontology/onet_additions.py for build_ontology.py to import."""
    out_path = Path(__file__).parent / "onet_additions.py"

    lines = [
        '"""',
        'Auto-generated from O*NET Technology Skills + explicit additions.',
        'Imported by build_ontology.py.',
        'Do not edit manually — re-run update_from_onet.py to regenerate.',
        '"""',
        '',
        'ONET_CHILD_OF: list[tuple[str, str]] = [',
    ]

    by_parent: dict[str, list[str]] = {}
    for skill, parent in additions:
        by_parent.setdefault(parent, []).append(skill)

    for parent in sorted(by_parent.keys()):
        lines.append(f"    # {parent}")
        for skill in sorted(by_parent[parent]):
            lines.append(f'    ("{skill}", "{parent}"),')

    lines += [']', '']

    out_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f"\nWrote {len(additions)} additions -> {out_path}")
    print("Next: python ontology/build_ontology.py")


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found.")
        print("Download Technology Skills Excel from onetcenter.org/database.html")
        return

    print(f"Reading O*NET Technology Skills from {XLSX_PATH.name}...")
    onet_pairs = extract_onet_skills(XLSX_PATH)
    print(f"Extracted {len(onet_pairs)} hot/in-demand tech skills from O*NET")

    additions = generate_additions(onet_pairs)
    print_report(additions)
    write_additions_file(additions)


if __name__ == "__main__":
    main()
